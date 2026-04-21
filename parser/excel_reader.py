import datetime
import uuid
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter

from .schema import CellNode, SheetNode, WorkbookNode
from .formula_parser import parse_formula_refs
from .section_detector import detect_sections


def _cell_value_type(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, datetime.datetime):
        return "date"
    return "string"


def _col_headers(ws, header_row: int) -> dict[str, str]:
    """Return {col_letter: header_value} for the given header row."""
    headers: dict[str, str] = {}
    for cell in ws[header_row]:
        if cell.value is not None and cell.column is not None:
            headers[get_column_letter(cell.column)] = str(cell.value)
    return headers


def _row_categories(ws, cat_col: str = "B") -> dict[int, str]:
    """Return {row: category_value} by scanning the category column."""
    cats: dict[int, str] = {}
    current = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.column is None or cell.row is None:
                continue
            if get_column_letter(cell.column) == cat_col:
                if cell.value and isinstance(cell.value, str):
                    current = cell.value.strip()
                if current:
                    cats[cell.row] = current
    return cats


def parse_sheet(wb_formula, wb_values, sheet_name: str,
                workbook_id: str, sheet_index: int,
                llm_provider=None) -> tuple[SheetNode, list[CellNode]]:
    """Parse a single sheet and return (SheetNode, list[CellNode]).

    wb_formula: workbook opened with data_only=False (has formulas)
    wb_values:  workbook opened with data_only=True  (has computed values)
    """
    ws_f = wb_formula[sheet_name]
    ws_v = wb_values[sheet_name]

    max_row = ws_f.max_row
    max_col = ws_f.max_column

    # Detect header row: first row that has ≥3 non-empty cells
    header_row = 3  # default for 参数输入表; generalised below
    for r in range(1, min(10, max_row + 1)):
        non_empty = sum(1 for c in ws_f[r] if c.value is not None)
        if non_empty >= 3:
            header_row = r
            break

    col_headers = _col_headers(ws_f, header_row)
    row_cats = _row_categories(ws_f, "B")

    # Collect merged cell info
    merged_ranges = []
    for mr in ws_f.merged_cells.ranges:
        merged_ranges.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))

    # Build rows_data for section detector
    rows_data: list[dict] = []
    for row in ws_f.iter_rows(min_row=header_row):
        row_num = row[0].row if row else header_row
        if row_num is None:
            continue
        row_dict: dict = {"_row": row_num, "_is_head": row_num == header_row}
        for cell in row:
            if cell.column is None:
                continue
            col_letter = get_column_letter(cell.column)
            row_dict[col_letter] = cell.value
        rows_data.append(row_dict)

    sections = detect_sections(
        sheet_name, rows_data, merged_ranges,
        llm_provider=llm_provider,
        sheet_meta={"max_row": max_row, "max_col": max_col},
    )

    # Build row → section_id lookup
    row_to_section: dict[int, str] = {}
    for sec in sections:
        if sec.row_start is not None and sec.row_end is not None:
            for r in range(sec.row_start, sec.row_end + 1):
                row_to_section[r] = sec.id

    # Parse cells
    cells: list[CellNode] = []
    for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
        for cell_f, cell_v in zip(row_f, row_v):
            if cell_f.value is None:
                continue

            row_num = cell_f.row
            col_num = cell_f.column
            if row_num is None or col_num is None:
                continue
            col_letter = get_column_letter(col_num)
            cell_id = f"{sheet_name}_{row_num}_{col_letter}"

            is_formula = isinstance(cell_f.value, str) and cell_f.value.startswith("=")
            formula_raw = cell_f.value if is_formula else None

            # Use computed value from data_only workbook; fall back to formula workbook value
            raw_value = cell_v.value if cell_v.value is not None else (None if is_formula else cell_f.value)

            formula_refs = parse_formula_refs(formula_raw, sheet_name) if formula_raw else []

            is_head = row_num == header_row

            # label: value in col D of same row (parameter name)
            d_cell = ws_f.cell(row=row_num, column=4)
            label = str(d_cell.value).strip() if d_cell.value and not is_head else None

            # description: value in col K of same row
            k_cell = ws_f.cell(row=row_num, column=11)
            description = str(k_cell.value).strip() if k_cell.value and not is_head else None

            # unit: value in col J of same row
            j_cell = ws_f.cell(row=row_num, column=10)
            unit = str(j_cell.value).strip() if j_cell.value and not is_head else None

            cells.append(CellNode(
                id=cell_id,
                sheet=sheet_name,
                row=row_num,
                col=col_letter,
                value=raw_value,
                value_type=_cell_value_type(raw_value),
                formula_raw=formula_raw,
                formula_refs=formula_refs,
                is_head=is_head,
                row_category=row_cats.get(row_num) if not is_head else None,
                col_category=col_headers.get(col_letter) if not is_head else None,
                unit=unit,
                label=label,
                description=description,
                section_id=row_to_section.get(row_num),
            ))

    sheet_node = SheetNode(
        id=sheet_name,
        workbook_id=workbook_id,
        index=sheet_index,
        row_count=max_row,
        col_count=max_col,
        sections=sections,
    )
    return sheet_node, cells


def parse_workbook(filepath: str,
                   sheet_names: Optional[list[str]] = None,
                   llm_provider=None) -> WorkbookNode:
    """Parse an Excel workbook and return a WorkbookNode with all cells.

    sheet_names: if provided, only parse those sheets; otherwise parse all.
    llm_provider: optional LLMProvider for intelligent section detection.
    """
    wb_formula = openpyxl.load_workbook(filepath, data_only=False)
    wb_values = openpyxl.load_workbook(filepath, data_only=True)

    workbook_id = str(uuid.uuid4())[:8]
    import os
    filename = os.path.basename(filepath)
    upload_time = datetime.datetime.now().isoformat()

    target_sheets = sheet_names or wb_formula.sheetnames

    workbook = WorkbookNode(
        id=workbook_id,
        filename=filename,
        upload_time=upload_time,
        sheet_count=len(target_sheets),
    )

    for idx, name in enumerate(target_sheets):
        if name not in wb_formula.sheetnames:
            continue
        sheet_node, cells = parse_sheet(wb_formula, wb_values, name, workbook_id, idx,
                                        llm_provider=llm_provider)
        sheet_node.sections  # already populated
        # Attach cells to sheet for downstream use
        sheet_node._cells = cells  # type: ignore[attr-defined]
        workbook.sheets.append(sheet_node)

    wb_formula.close()
    wb_values.close()
    return workbook
